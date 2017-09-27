import openpyxl, time, os
from openpyxl import load_workbook

#introduction
print('What is the first file you would like to add?')
first = int(input())
print('Thanks! Stored.')
print()

print('What is the last report you would like to add through?')
last = int(input())
print('Thanks! Stored.')
print('Running...')
print('Time stamp: ' + time.strftime('%H:%M:%S EST on %b %d, %Y'))

#master workbook
mwb = openpyxl.load_workbook('masterFile.xlsx')
mws = mwb.active

#workbook to copy from
fileNumber = int(first)
fileName = 'report' + str(fileNumber) + '.xlsx'
wb = openpyxl.load_workbook(fileName)
ws = wb.active

lastRow = int(ws.max_row)
lastCellName = str('J' + str(lastRow))


#actual program loop
for i in range (first, (last + 1)):
    
    wb = openpyxl.load_workbook(fileName)
    ws = wb.active

    lastRow = int(ws.max_row)
    lastCellName = str('J' + str(lastRow))
    
    for rowOfCellObjects in ws['A2':lastCellName]:
        cellList = []
        for cellObj in rowOfCellObjects:
            cellList.append(cellObj.value)
        mws.append(cellList)

    mwb.save('masterTest.xlsx')
    fileNumber = int(fileNumber) + 1
    fileName = 'report' + str(fileNumber) + '.xlsx'
    
    
#save and complete
mwb.save('masterTest.xlsx')
print()
print()
print('Complete!')
print('Time stamp: ' + time.strftime('%H:%M:%S EST on %b %d, %Y'))
